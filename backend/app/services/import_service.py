"""Import service for parsing and importing ROPA records from Thai Excel form."""

from datetime import date, datetime
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.controller import Controller
from app.models.data_subject_category import DataSubjectCategory
from app.models.department import Department
from app.models.import_batch import ImportBatch
from app.models.personal_data_type import PersonalDataType
from app.models.processor import Processor
from app.models.record_version import RecordVersion
from app.models.ropa_data_subject import RopaDataSubject
from app.models.ropa_personal_data_type import RopaPersonalDataType
from app.models.ropa_record import RopaRecord
from app.models.user import User
from app.schemas.import_export import ImportPreviewResponse, ImportRowData, ImportRowError
from app.services.audit_service import log_action


# Sheet-specific Thai column indices (1-based)

PROCESSOR_COLUMN_MAP = {
    # Section 0: Metadata & Row ID
    "seq": 1,  # A

    # Section 1: Identity
    "processor_name": 2,   # B ชื่อผู้ประมวลผลข้อมูลส่วนบุคคล
    "controller_address": 3,  # C ที่อยู่ผู้ควบคุมข้อมูลส่วนบุคคล

    # Section 2: Activity & Purpose
    "activity_name": 4,    # D
    "purpose": 5,          # E

    # Section 3: Data Categories
    "personal_data_types": 6,       # F
    "data_subject_categories": 7,   # G
    "data_type_general": 8,         # H

    # Section 4: Data Source
    "data_acquisition_method": 9,   # I
    "data_source_direct": 10,       # J
    "data_source_other": 11,        # K

    # Section 5: Legal Basis
    "legal_basis_thai": 12,         # L

    # Section 6: Cross-border Transfer
    "cross_border_transfer": 13,    # M
    "cross_border_affiliate": 14,   # N
    "cross_border_method": 15,      # O
    "cross_border_standard": 16,    # P
    "cross_border_exception": 17,   # Q

    # Section 7: Retention & Storage
    "storage_type": 18,             # R
    "storage_method": 19,           # S
    "retention_period": 20,         # T
    "access_rights": 21,            # U
    "deletion_method": 22,          # V

    # Section 8: Security Measures
    "security_organizational": 23,  # W
    "security_technical": 24,       # X
    "security_physical": 25,        # Y
    "security_access_control": 26,  # Z
    "security_responsibility": 27,  # AA
    "security_audit": 28,           # AB
}

# ใช้สำหรับ Example / Controller-like structure
CONTROLLER_COLUMN_MAP = {
    "seq": 1,                       # A
    "controller_name": 2,           # B
    "activity_name": 3,             # C
    "purpose": 4,                   # D
    "personal_data_types": 5,       # E
    "data_subject_categories": 6,   # F
    "data_type_general": 7,         # G
    "data_acquisition_method": 8,   # H
    "data_source_direct": 9,        # I
    "data_source_other": 10,        # J
    "legal_basis_thai": 11,         # K
    "minor_consent_under_10": 12,   # L
    "minor_consent_10_20": 13,      # M
    "cross_border_transfer": 14,    # N
    "cross_border_affiliate": 15,   # O
    "cross_border_method": 16,      # P
    "cross_border_standard": 17,    # Q
    "cross_border_exception": 18,   # R
    "storage_type": 19,             # S
    "storage_method": 20,           # T
    "retention_period": 21,         # U
    "access_rights": 22,            # V
    "deletion_method": 23,          # W
    "disclosure_exemption": 24,     # X
    "rights_refusal": 25,           # Y
    "security_organizational": 26,  # Z
    "security_technical": 27,       # AA
    "security_physical": 28,        # AB
    "security_access_control": 29,  # AC
    "security_responsibility": 30,  # AD
    "security_audit": 31,           # AE
}

EXAMPLE_CONTROLLER_COLUMN_MAP = {
    "seq": 1,                       # A
    "controller_name": 2,           # B
    "activity_name": 3,             # C
    "purpose": 4,                   # D
    "personal_data_types": 5,       # E
    "data_subject_categories": 6,   # F
    "data_type_general": 7,         # G
    "data_acquisition_method": 8,   # H
    "data_source_direct": 9,        # I
    "data_source_other": 10,        # J
    "legal_basis_thai": 11,         # K
    "minor_consent_under_10": 12,   # L
    "minor_consent_10_20": 13,      # M
    "cross_border_transfer": 14,    # N
    "cross_border_affiliate": 15,   # O
    "cross_border_method": 16,      # P
    "cross_border_standard": 17,    # Q
    "cross_border_exception": 18,   # R
    "storage_type": 19,             # S
    "storage_method": 20,           # T
    "retention_period": 21,         # U
    "access_rights": 22,            # V
    "deletion_method": 23,          # W
    "disclosure_exemption": 24,     # X
    "rights_refusal": 25,           # Y
    "security_organizational": 26,  # Z
    "security_technical": 27,       # AA
    "security_physical": 28,        # AB
    "security_access_control": 29,  # AC
    "security_responsibility": 30,  # AD
    "security_audit": 31,           # AE
}

def _should_skip_sheet(ws: Worksheet) -> bool:
    sheet_name = ws.title.lower().strip()

    # skip only real helper/instruction sheets if you have any
    return sheet_name in {"instruction", "instructions", "คำอธิบาย"}

def _detect_sheet_type(ws: Worksheet) -> str:
    """
    Detect whether a sheet is Controller or Processor.
    """
    sheet_name = ws.title.lower().strip()

    if "processor" in sheet_name:
        return "Processor"
    return "Controller"

def _detect_sheet_kind(ws: Worksheet) -> str:
    """
    Return one of:
    - processor
    - controller
    - example_controller
    """
    sheet_name = ws.title.lower().strip()

    if "processor" in sheet_name:
        return "processor"

    if "example" in sheet_name or "ตัวอย่าง" in sheet_name:
        return "example_controller"

    return "controller"

def _sheet_kind_to_role_type(sheet_kind: str) -> str:
    if sheet_kind == "processor":
        return "Processor"
    return "Controller"

def _get_column_map(sheet_kind: str) -> dict:
    if sheet_kind == "processor":
        return PROCESSOR_COLUMN_MAP

    if sheet_kind == "example_controller":
        return EXAMPLE_CONTROLLER_COLUMN_MAP

    return EXAMPLE_CONTROLLER_COLUMN_MAP

def _get_cell_value(row, col_idx: int) -> Optional[str]:
    """Get cell value at 1-based column index (col_idx: 1 = column A, 2 = column B)."""
    if col_idx > len(row):
        return None
    cell = row[col_idx - 1]  # Convert to 0-based index
    return _cell_str(cell.value)



def _parse_bool(value: Optional[str]) -> Optional[bool]:
    """Parse Y/N/Yes/No/True/False to boolean."""
    if value is None:
        return None
    v = str(value).strip().lower()
    if v in ("y", "yes", "true", "1"):
        return True
    if v in ("n", "no", "false", "0"):
        return False
    return None


def _parse_date(value) -> Optional[date]:
    """Parse a date value from Excel (could be datetime, date, or string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Try parsing string formats
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_str(value) -> Optional[str]:
    """Convert a cell value to a stripped string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_name(name: Optional[str]) -> str:
    """Normalize a name for matching: strip whitespace, lowercase, remove extra spaces."""
    if not name:
        return ""
    # Remove leading/trailing whitespace, convert to lowercase, collapse multiple spaces
    return " ".join(str(name).strip().lower().split())


def _build_lookup_maps(db: Session) -> dict:
    """Build lookup maps for departments, controllers, processors, data subjects, personal data types."""
    departments = db.query(Department).filter(Department.is_active == True).all()
    dept_by_name = {_normalize_name(d.name): d for d in departments}
    dept_by_code = {d.code.lower(): d for d in departments}

    controllers = db.query(Controller).filter(Controller.is_active == True).all()
    ctrl_by_name = {_normalize_name(c.name): c for c in controllers}

    processors = db.query(Processor).filter(Processor.is_active == True).all()
    proc_by_name = {_normalize_name(p.name): p for p in processors}

    ds_categories = db.query(DataSubjectCategory).all()
    ds_by_name = {d.name.lower(): d for d in ds_categories}

    pdt_types = db.query(PersonalDataType).all()
    pdt_by_name = {p.name.lower(): p for p in pdt_types}

    return {
        "dept_by_name": dept_by_name,
        "dept_by_code": dept_by_code,
        "ctrl_by_name": ctrl_by_name,
        "proc_by_name": proc_by_name,
        "ds_by_name": ds_by_name,
        "pdt_by_name": pdt_by_name,
    }


def _validate_row(
        row,
        row_number: int,
        sheet_name: str,
        sheet_kind: str,
        role_type: str,
        lookups: dict,
        inherited_controller: Optional[str] = None,
        inherited_activity: Optional[str] = None,
    ) -> tuple[Optional[ImportRowData], list[ImportRowError]]:
    errors: list[ImportRowError] = []
    col_map = _get_column_map(sheet_kind)

    def get_col(field_name: str) -> Optional[str]:
        col_idx = col_map.get(field_name)
        if not col_idx:
            return None
        if col_idx <= len(row):
            return _get_cell_value(row, col_idx)
        return None

    # Identity
    controller_name = get_col("controller_name")
    processor_name = get_col("processor_name")
    activity_name = get_col("activity_name")

    # For continuation rows
    inherited_name = inherited_controller
    current_name = processor_name if role_type == "Processor" else controller_name

    if not activity_name and inherited_activity:
        activity_name = inherited_activity
    if not current_name and inherited_name:
        current_name = inherited_name

    if role_type == "Processor":
        processor_name = current_name
    else:
        controller_name = current_name

    # Core fields
    purpose = get_col("purpose")
    personal_data_desc = get_col("personal_data_types")
    data_subject_desc = get_col("data_subject_categories")
    data_type_general = get_col("data_type_general")
    data_acq_method = get_col("data_acquisition_method")
    data_src_direct = get_col("data_source_direct")
    data_src_other = get_col("data_source_other")
    legal_basis = get_col("legal_basis_thai")
    legal_basis_gdpr = get_col("legal_basis_gdpr")

    minor_consent_under_10 = get_col("minor_consent_under_10")
    minor_consent_10_20 = get_col("minor_consent_10_20")

    cross_border_str = get_col("cross_border_transfer")
    cross_border_affiliate = get_col("cross_border_affiliate")
    cross_border_method = get_col("cross_border_method")
    cross_border_std = get_col("cross_border_standard")
    cross_border_exc = get_col("cross_border_exception")

    storage_type = get_col("storage_type")
    storage_method = get_col("storage_method")
    retention_period = get_col("retention_period")
    access_rights = get_col("access_rights")
    deletion_method = get_col("deletion_method")

    data_owner = get_col("data_owner")
    third_party_recipients = get_col("third_party_recipients")
    disclosure_exemption = get_col("disclosure_exemption")
    rights_refusal = get_col("rights_refusal")

    security_organizational = get_col("security_organizational")
    security_technical = get_col("security_technical")
    security_physical = get_col("security_physical")
    security_access_control = get_col("security_access_control")
    security_responsibility = get_col("security_responsibility")
    security_audit = get_col("security_audit")

    excel_address = get_col("controller_address") or get_col("address")

    # Validation
    effective_name = processor_name if role_type == "Processor" else controller_name

    if not effective_name and not activity_name:
        errors.append(
            ImportRowError(
                sheet_name=sheet_name,
                row_number=row_number,
                field_name="row_data",
                error_reason="Row must have either entity name or activity name",
            )
        )

    department_id = None

    # Parse cross-border boolean
    cross_border_transfer = None
    if cross_border_str:
        raw = str(cross_border_str).strip()
        v = raw.lower()

        false_tokens = ["ไม่มี", "ไม่", "no", "n", "false", "0"]
        true_tokens = ["มี", "yes", "y", "true", "1", "✓", "ü"]

        if any(token in raw for token in ["ไม่มี", "ไม่"]) or v in ["no", "n", "false", "0"]:
            cross_border_transfer = False
        elif any(token in raw for token in ["มี", "✓", "ü"]) or v in ["yes", "y", "true", "1"]:
            cross_border_transfer = True
            
    controller_id = None
    processor_id = None
    controller_name_stored = None
    processor_name_stored = None

    controller_address = None
    controller_email = None
    controller_phone = None
    processor_address = None
    processor_email = None
    processor_phone = None

    if role_type == "Controller" and controller_name:
        ctrl = lookups["ctrl_by_name"].get(_normalize_name(controller_name))
        if ctrl:
            controller_id = ctrl.id
            controller_address = ctrl.address
            controller_email = ctrl.email
            controller_phone = ctrl.phone
        controller_name_stored = controller_name

    elif role_type == "Processor" and processor_name:
        proc = lookups["proc_by_name"].get(_normalize_name(processor_name))
        if proc:
            processor_id = proc.id
            processor_address = proc.address
            processor_email = proc.email
            processor_phone = proc.phone
        processor_name_stored = processor_name

    # Parse data subject categories
    ds_ids: list[int] = []
    if data_subject_desc:
        categories = ["พนักงาน", "ลูกค้า", "คู่ค้า", "ผู้ติดต่อ", "ผู้สมัคร", "สมาชิก", "อื่น"]
        for cat in categories:
            if cat in data_subject_desc:
                ds = lookups["ds_by_name"].get(cat.lower())
                if ds:
                    ds_ids.append(ds.id)

    # Parse personal data types
    pdt_ids: list[int] = []
    if personal_data_desc:
        pdt_keywords = [
            "ชื่อ", "นามสกุล", "เบอร์โทร", "อีเมล", "ที่อยู่", "เลขประจำตัว",
            "ข้อมูลธนาคาร", "วันเดือนปี", "บัญชี", "บัตร", "รหัส", "สุขภาพ",
            "ศาสนา", "ประวัติ", "ภาพถ่าย", "ลายนิ้วมือ"
        ]
        for keyword in pdt_keywords:
            if keyword in personal_data_desc:
                pdt = lookups["pdt_by_name"].get(keyword.lower())
                if pdt:
                    pdt_ids.append(pdt.id)

    if errors:
        return None, errors

    row_data = ImportRowData(
        sheet_name=sheet_name,
        row_number=row_number,
        role_type=role_type,
        department_id=department_id,
        controller_id=controller_id,
        processor_id=processor_id,
        controller_name=controller_name_stored,
        processor_name=processor_name_stored,

        excel_address=excel_address,
        excel_personal_data_types=personal_data_desc,
        excel_data_subject_categories=data_subject_desc,
        excel_data_type_general=data_type_general,

        controller_address=controller_address,
        controller_email=controller_email,
        controller_phone=controller_phone,
        processor_address=processor_address,
        processor_email=processor_email,
        processor_phone=processor_phone,

        data_subject_category_ids=ds_ids,
        personal_data_type_ids=pdt_ids,

        activity_name=activity_name,
        purpose=purpose,
        risk_level=None,

        data_acquisition_method=data_acq_method,
        data_source_direct=data_src_direct,
        data_source_other=data_src_other,

        legal_basis_thai=legal_basis,
        legal_basis_gdpr=legal_basis_gdpr,

        minor_consent_under_10=minor_consent_under_10,
        minor_consent_10_20=minor_consent_10_20,

        cross_border_transfer=cross_border_transfer,
        cross_border_affiliate=cross_border_affiliate,
        cross_border_method=cross_border_method,
        cross_border_standard=cross_border_std,
        cross_border_exception=cross_border_exc,

        retention_period=retention_period,
        storage_type=storage_type,
        storage_method=storage_method,
        access_rights=access_rights,
        deletion_method=deletion_method,

        data_owner=data_owner,
        third_party_recipients=third_party_recipients,
        disclosure_exemption=disclosure_exemption,
        rights_refusal=rights_refusal,

        security_organizational=security_organizational,
        security_technical=security_technical,
        security_physical=security_physical,
        security_access_control=security_access_control,
        security_responsibility=security_responsibility,
        security_audit=security_audit,
    )
    return row_data, []

def _parse_sheet(
    ws: Worksheet,
    sheet_name: str,
    lookups: dict,
) -> tuple[list[ImportRowData], list[ImportRowError], int]:
    if _should_skip_sheet(ws):
        return [], [], 0

    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        return [], [], 0

    sheet_kind = _detect_sheet_kind(ws)
    role_type = _sheet_kind_to_role_type(sheet_kind)
    col_map = _get_column_map(sheet_kind)

    valid_rows: list[ImportRowData] = []
    all_errors: list[ImportRowError] = []
    rows_with_data = 0

    # Find first data row by looking for numeric sequence in first column
    first_data_row = None
    for row_idx, row in enumerate(rows):
        if row_idx < 3:
            continue

        seq_cell = row[0] if len(row) > 0 else None
        if seq_cell and seq_cell.value is not None:
            try:
                int(seq_cell.value)
                first_data_row = row_idx
                break
            except (ValueError, TypeError):
                continue

    if first_data_row is None:
        first_data_row = 4

    prev_name = None
    prev_activity = None

    for row_idx, row in enumerate(rows[first_data_row:], start=first_data_row + 1):
        name_field = "processor_name" if role_type == "Processor" else "controller_name"

        activity_val = _get_cell_value(row, col_map["activity_name"]) if col_map.get("activity_name") else None
        name_val = _get_cell_value(row, col_map[name_field]) if col_map.get(name_field) else None
        personal_data_val = _get_cell_value(row, col_map["personal_data_types"]) if col_map.get("personal_data_types") else None

        if not personal_data_val and not name_val and not activity_val:
            continue

        effective_activity = activity_val or prev_activity
        effective_name = name_val or prev_name

        if not effective_name and not effective_activity and not personal_data_val:
            continue

        rows_with_data += 1

        parsed, errors = _validate_row(
            row,
            row_idx,
            sheet_name,
            sheet_kind,
            role_type,
            lookups,
            inherited_controller=prev_name if not name_val else None,
            inherited_activity=prev_activity if not activity_val else None,
        )

        if parsed:
            valid_rows.append(parsed)
            current_name = parsed.processor_name if role_type == "Processor" else parsed.controller_name
            prev_name = current_name or prev_name
            prev_activity = parsed.activity_name or prev_activity

        all_errors.extend(errors)

    return valid_rows, all_errors, rows_with_data


def _check_duplicate_record(db: Session, row_data: ImportRowData) -> tuple[bool, Optional[int]]:
    """
    Check if the imported row data matches an existing ROPA record.
    Only checks after we have confirmed controller/processor IDs.
    
    Returns:
        (is_duplicate: bool, duplicate_record_id: Optional[int])
    
    Matches on:
    - Same role_type
    - Same controller_id/processor_id (must be matched)
    - Same or very similar activity_name
    """
    # Only check if we have confirmed controller/processor ID
    if not row_data.controller_id and not row_data.processor_id:
        return False, None
    
    # Need activity name for reliable duplicate detection
    if not row_data.activity_name:
        return False, None
    
    query = db.query(RopaRecord).filter(
        RopaRecord.is_deleted == False,
        RopaRecord.role_type == row_data.role_type,
    )
    
    # Filter by controller/processor
    if row_data.role_type == "Controller" and row_data.controller_id:
        query = query.filter(RopaRecord.controller_id == row_data.controller_id)
    elif row_data.role_type == "Processor" and row_data.processor_id:
        query = query.filter(RopaRecord.processor_id == row_data.processor_id)
    else:
        return False, None
    
    # Try exact activity name match first
    activity_name_normalized = _normalize_name(row_data.activity_name)
    
    # Try to find by normalized activity name
    for record in query.all():
        existing_normalized = _normalize_name(record.activity_name) if record.activity_name else ""
        if activity_name_normalized and existing_normalized == activity_name_normalized:
            return True, record.id
    
    # If no exact match, don't mark as duplicate (to be safe)
    return False, None


def preview_import(db: Session, file_content: bytes) -> ImportPreviewResponse:
    """Parse a Thai ROPA form Excel file and return a preview with valid rows and errors."""
    from app.schemas.import_export import ControllerProcessorOption
    
    wb = load_workbook(filename=BytesIO(file_content), read_only=False, data_only=True)
    lookups = _build_lookup_maps(db)

    all_valid: list[ImportRowData] = []
    all_errors: list[ImportRowError] = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        valid, errors, sheet_row_count = _parse_sheet(ws, sheet_name, lookups)
        all_valid.extend(valid)
        all_errors.extend(errors)
        total_rows += sheet_row_count

    wb.close()

    # Check for duplicates in valid rows (only if controller/processor ID is already matched)
    for row in all_valid:
        # Only check if we have confirmed IDs (will not be auto-created)
        if row.controller_id or row.processor_id:
            is_duplicate, dup_id = _check_duplicate_record(db, row)
            if is_duplicate:
                row.is_duplicate = True
                row.duplicate_record_id = dup_id

    # Build controller and processor options
    controller_options = [
        ControllerProcessorOption(
            id=c.id, 
            name=c.name, 
            type="controller",
            is_active=c.is_active,
            address=c.address,
            email=c.email,
            phone=c.phone
        )
        for c in lookups["ctrl_by_name"].values()
    ]
    
    processor_options = [
        ControllerProcessorOption(
            id=p.id, 
            name=p.name, 
            type="processor",
            is_active=p.is_active,
            address=p.address,
            email=p.email,
            phone=p.phone
        )
        for p in lookups["proc_by_name"].values()
    ]

    return ImportPreviewResponse(
        valid_rows=all_valid,
        errors=all_errors,
        total_rows=total_rows,
        valid_count=len(all_valid),
        error_count=len(all_errors),
        controller_options=controller_options,
        processor_options=processor_options,
    )


def _get_or_create_controller(db: Session, name: str, address: Optional[str], email: Optional[str], phone: Optional[str], user_id: int) -> int:
    """Get controller by name or create a new one if it doesn't exist."""
    normalized_name = _normalize_name(name)
    
    # Try to find existing controller by normalized name
    existing = db.query(Controller).filter(
        Controller.name.ilike(f"%{name}%")
    ).first()
    
    if existing:
        return existing.id
    
    # Create new controller with data from import
    new_controller = Controller(
        name=name,
        address=address,
        email=email,
        phone=phone,
        is_active=True
    )
    db.add(new_controller)
    db.flush()
    
    log_action(
        db,
        user_id=user_id,
        action="create",
        table_name="controllers",
        record_id=new_controller.id,
        new_value={"name": name, "address": address, "email": email, "phone": phone},
        reason="Auto-created during Excel import"
    )
    
    return new_controller.id


def _get_or_create_processor(db: Session, name: str, address: Optional[str], email: Optional[str], phone: Optional[str], user_id: int, source_controller_id: Optional[int] = None) -> int:
    """Get processor by name or create a new one if it doesn't exist."""
    normalized_name = _normalize_name(name)
    
    # Try to find existing processor by normalized name
    existing = db.query(Processor).filter(
        Processor.name.ilike(f"%{name}%")
    ).first()
    
    if existing:
        return existing.id
    
    # Create new processor with data from import
    new_processor = Processor(
        name=name,
        address=address,
        email=email,
        phone=phone,
        source_controller_id=source_controller_id,
        is_active=True
    )
    db.add(new_processor)
    db.flush()
    
    log_action(
        db,
        user_id=user_id,
        action="create",
        table_name="processors",
        record_id=new_processor.id,
        new_value={"name": name, "address": address, "email": email, "phone": phone, "source_controller_id": source_controller_id},
        reason="Auto-created during Excel import"
    )
    
    return new_processor.id


def confirm_import(
    db: Session,
    file_content: bytes,
    filename: str,
    user_id: int,
    target_department_id: Optional[int] = None,
    row_mappings: Optional[dict[str, int]] = None,
) -> ImportBatch:
    """
    Re-parse the file and import only valid rows, creating ROPA records.
    
    Args:
        db: Database session
        file_content: Excel file bytes
        filename: Original filename
        user_id: User performing import
        target_department_id: Explicit department ID to use (overrides inference)
        row_mappings: Dict mapping row_keys (sheet-row) to controller/processor IDs.
                      If provided, uses these IDs instead of Excel name matching.
                      Example: {"Controller-15": 5, "Processor-20": 10}
    
    Raises:
        ValueError: If department_id cannot be determined for import
    """
    wb = load_workbook(filename=BytesIO(file_content), read_only=False, data_only=True)
    lookups = _build_lookup_maps(db)

    all_valid: list[ImportRowData] = []
    all_errors: list[ImportRowError] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        valid, errors, _ = _parse_sheet(ws, sheet_name, lookups)
        all_valid.extend(valid)
        all_errors.extend(errors)

    wb.close()

    rows_success = 0

    # Determine default department for rows without one
    default_dept_id = None
    
    # Priority 1: Explicit target department
    if target_department_id:
        default_dept_id = target_department_id
    # Priority 2: Current user's department
    else:
        current_user = db.query(User).filter(User.id == user_id).first()
        if current_user and current_user.department_id:
            default_dept_id = current_user.department_id
    
    # Ensure we have a department to use
    if not default_dept_id:
        raise ValueError(
            "ไม่สามารถระบุแผนกสำหรับการนำเข้าได้\n\n"
            "กรุณาระบุ department_id ในการร้องขอ หรือ\n"
            "ตั้งค่า department ให้กับผู้ใช้ที่ทำการนำเข้า\n\n"
            "คำแนะนำ: POST /api/import/confirm?department_id=1 -F file=@form.xlsx"
        )

    for row in all_valid:
        # Use row's department if available, otherwise use default
        dept_id = row.department_id or default_dept_id
        
        # Handle controller/processor ID assignment with auto-creation if needed
        controller_id = row.controller_id
        processor_id = row.processor_id
        
        # For Controller: if no match was found but we have a name, create new controller
        if row.role_type == "Controller" and controller_id is None and row.controller_name:
            controller_id = _get_or_create_controller(
                db, 
                row.controller_name, 
                row.controller_address or row.excel_address,
                row.controller_email,
                row.controller_phone,
                user_id
            )
        
        # For Processor: if no match was found but we have a name, create new processor
        if row.role_type == "Processor" and processor_id is None and row.processor_name:
            processor_id = _get_or_create_processor(
                db,
                row.processor_name,
                row.processor_address or row.excel_address,
                row.processor_email,
                row.processor_phone,
                user_id,
                source_controller_id=None  # No specific source controller for imported processors
            )

        # Check for duplicate after controller/processor is assigned
        row_for_dup_check = ImportRowData(
            sheet_name=row.sheet_name,
            row_number=row.row_number,
            role_type=row.role_type,
            controller_id=controller_id,
            processor_id=processor_id,
            activity_name=row.activity_name,
            purpose=row.purpose,
            controller_name=row.controller_name,
            processor_name=row.processor_name,
        )
        is_duplicate, dup_id = _check_duplicate_record(db, row_for_dup_check)
        if is_duplicate:
            # Skip this row - it's a duplicate
            all_errors.append(ImportRowError(
                sheet_name=row.sheet_name,
                row_number=row.row_number,
                field_name="row_data",
                error_reason=f"แถวนี้ซ้ำกับ ROPA Record #{dup_id} ที่มีอยู่แล้ว"
            ))
            continue

        record = RopaRecord(
            department_id=dept_id,
            created_by=user_id,
            role_type=row.role_type,
            status="pending_approval",
            is_deleted=False,
            controller_id=controller_id,
            processor_id=processor_id,

            # Section 1: Basic Information
            activity_name=row.activity_name,
            purpose=row.purpose,
            risk_level=row.risk_level,

            # Section 2: Data Source
            data_acquisition_method=row.data_acquisition_method,
            data_source_direct=row.data_source_direct,
            data_source_other=row.data_source_other,

            # Section 3: Legal Basis
            legal_basis_thai=row.legal_basis_thai,

            # Section 4: Minor Consent
            minor_consent_under_10=row.minor_consent_under_10,
            minor_consent_10_20=row.minor_consent_10_20,

            # Section 5: Cross-Border Transfer
            cross_border_transfer=row.cross_border_transfer,
            cross_border_affiliate=row.cross_border_affiliate,
            cross_border_method=row.cross_border_method,
            cross_border_standard=row.cross_border_standard,
            cross_border_exception=row.cross_border_exception,

            # Section 6: Retention Policy
            retention_period=row.retention_period,
            retention_expiry_date=None,
            next_review_date=None,
            storage_type=row.storage_type,
            storage_method=row.storage_method,
            access_rights=row.access_rights,
            deletion_method=row.deletion_method,

            # Section 7: Data Usage/Disclosure
            data_owner=row.data_owner,
            third_party_recipients=row.third_party_recipients,
            disclosure_exemption=row.disclosure_exemption,
            rights_refusal=row.rights_refusal,

            # Section 8: Security Measures
            security_organizational=row.security_organizational,
            security_technical=row.security_technical,
            security_physical=row.security_physical,
            security_access_control=row.security_access_control,
            security_responsibility=row.security_responsibility,
            security_audit=row.security_audit,
        )
        db.add(record)
        db.flush()

        # Link many-to-many relationships (deduplicate IDs)
        for ds_id in set(row.data_subject_category_ids):
            db.add(RopaDataSubject(ropa_record_id=record.id, data_subject_category_id=ds_id))
        
        for pdt_id in set(row.personal_data_type_ids):
            db.add(RopaPersonalDataType(ropa_record_id=record.id, personal_data_type_id=pdt_id))

        # Create initial version record (version 1)
        db.add(RecordVersion(
            ropa_record_id=record.id,
            version_number=1,
            changed_by=user_id,
            change_reason="Imported from Excel file",
            snapshot={
                "activity_name": record.activity_name,
                "purpose": record.purpose,
                "role_type": record.role_type,
                "status": record.status,
            },
        ))

        rows_success += 1

    # Commit all changes
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise e

    # Final error count includes both parsing errors and duplicate detection errors
    rows_failed = len(all_errors)

    # Create ImportBatch record
    error_details = None
    if all_errors:
        error_details = {"errors": [e.model_dump() for e in all_errors]}

    batch_status = "completed"
    if rows_success == 0 and rows_failed > 0:
        batch_status = "failed"
    elif rows_failed > 0:
        batch_status = "partial"

    batch = ImportBatch(
        imported_by=user_id,
        filename=filename,
        rows_success=rows_success,
        rows_failed=rows_failed,
        status=batch_status,
        error_details=error_details,
    )
    db.add(batch)
    db.flush()

    # Audit log
    log_action(
        db,
        user_id=user_id,
        action="import",
        table_name="ropa_records",
        record_id=batch.id,
        new_value={
            "filename": filename,
            "rows_success": rows_success,
            "rows_failed": rows_failed,
            "status": batch_status,
        },
        reason=f"Imported {rows_success} ROPA records from {filename}" + 
               (f" ({rows_failed} errors)" if rows_failed > 0 else ""),
    )

    db.commit()
    return batch


def list_import_batches(db: Session, page: int = 1, per_page: int = 20) -> dict:
    """List import batches with pagination."""
    query = db.query(ImportBatch).order_by(ImportBatch.created_at.desc())
    total = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()
    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page else 0,
    }
