"""Export service — generates Excel matching the official ROPA Form format."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.models.ropa_record import RopaRecord
from app.models.processor import Processor
from app.services.audit_service import log_action

# ─── Styles ────────────────────────────────────────────────────────────────────

# Dark blue — main group headers (row 1)
DARK_BLUE = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
# Medium blue — sub-headers (row 2)
MED_BLUE  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
# Light blue — section title row
LIGHT_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

WHITE_BOLD  = Font(bold=True, color="FFFFFF", name="TH SarabunPSK", size=11)
DARK_BOLD   = Font(bold=True, color="1F3864", name="TH SarabunPSK", size=11)
NORMAL_FONT = Font(name="TH SarabunPSK", size=11)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _style(cell, font=None, fill=None, alignment=None, border=None):
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ─── Controller columns ────────────────────────────────────────────────────────
# Each entry: (group_label, sub_label, field, width)
# group_label=None → merge with previous group cell
CONTROLLER_COLS = [
    # group,                                          sub,                                                    field,                      w
    ("ลำดับ",                                          None,                                                   "__seq__",                  6),
    ("1. ชื่อเจ้าของข้อมูลส่วนบุคคล",                  None,                                                   "data_subject_categories",  28),
    ("2. กิจกรรมประมวลผล",                             None,                                                   "activity_name",            28),
    ("3. วัตถุประสงค์ของการประมวลผล",                  None,                                                   "purpose",                  28),
    ("4. ข้อมูลส่วนบุคคลที่จัดเก็บ",                   None,                                                   "personal_data_types",      28),
    ("5. หมวดหมู่ของข้อมูล",                           None,                                                   "data_subject_categories",  22),
    ("6. ประเภทของข้อมูล",                             None,                                                   "personal_data_types",      22),
    ("7. วิธีการได้มาซึ่งข้อมูล",                      None,                                                   "data_acquisition_method",  22),
    ("8. แหล่งที่ได้มาซึ่งข้อมูล",                     "จากเจ้าของข้อมูลโดยตรง",                               "data_source_direct",       22),
    (None,                                             "จากแหล่งอื่น",                                         "data_source_other",        22),
    ("9. ฐานในการประมวลผล",                            None,                                                   "legal_basis_thai",         28),
    ("10. การขอความยินยอมของผู้เยาว์",                  "อายุไม่เกิน 10 ปี",                                    "minor_consent_under_10",   18),
    (None,                                             "อายุ 10-20 ปี",                                        "minor_consent_10_20",      18),
    ("11. ส่งหรือโอนข้อมูลไปต่างประเทศ",               "มีการส่งหรือโอนหรือไม่",                               "cross_border_transfer",    18),
    (None,                                             "บริษัทในเครือ",                                        "cross_border_affiliate",   22),
    (None,                                             "วิธีการโอนข้อมูล",                                     "cross_border_method",      22),
    (None,                                             "มาตรฐานการคุ้มครองข้อมูล",                             "cross_border_standard",    22),
    (None,                                             "ข้อยกเว้นตามมาตรา 28",                                 "cross_border_exception",   28),
    ("12. นโยบายการเก็บรักษาข้อมูล",                   "ประเภทของข้อมูลที่จัดเก็บ",                            "storage_type",             18),
    (None,                                             "วิธีการเก็บรักษา",                                     "storage_method",           22),
    (None,                                             "ระยะเวลาการเก็บรักษา",                                 "retention_period",         18),
    (None,                                             "สิทธิและวิธีการเข้าถึงข้อมูล",                         "access_rights",            28),
    (None,                                             "วิธีการลบหรือทำลายข้อมูล",                             "deletion_method",          22),
    ("13. การใช้หรือเปิดเผยที่ได้รับยกเว้น",            None,                                                   "disclosure_exemption",     28),
    ("14. การปฏิเสธคำขอใช้สิทธิ",                      None,                                                   "rights_refusal",           28),
    ("15. มาตรการรักษาความมั่นคงปลอดภัย",              "มาตรการเชิงองค์กร",                                    "security_organizational",  22),
    (None,                                             "มาตรการเชิงเทคนิค",                                    "security_technical",       22),
    (None,                                             "มาตรการทางกายภาพ",                                     "security_physical",        22),
    (None,                                             "การควบคุมการเข้าถึงข้อมูล",                            "security_access_control",  22),
    (None,                                             "การกำหนดหน้าที่ความรับผิดชอบ",                         "security_responsibility",  22),
    (None,                                             "มาตรการตรวจสอบย้อนหลัง",                               "security_audit",           22),
]

# ─── Processor columns ─────────────────────────────────────────────────────────
PROCESSOR_COLS = [
    ("ลำดับ",                                          None,                                                   "__seq__",                  6),
    ("1. ชื่อผู้ประมวลผลข้อมูลส่วนบุคคล",              None,                                                   "processor_name",           28),
    ("2. ที่อยู่ผู้ควบคุมข้อมูลส่วนบุคคล",             None,                                                   "source_controller",        28),
    ("3. กิจกรรมประมวลผล",                             None,                                                   "activity_name",            28),
    ("4. วัตถุประสงค์ของการประมวลผล",                  None,                                                   "purpose",                  28),
    ("5. ข้อมูลส่วนบุคคลที่จัดเก็บ",                   None,                                                   "personal_data_types",      28),
    ("6. หมวดหมู่ของข้อมูล",                           None,                                                   "data_subject_categories",  22),
    ("7. ประเภทของข้อมูล",                             None,                                                   "data_category",            22),
    ("8. วิธีการได้มาซึ่งข้อมูล",                      None,                                                   "data_acquisition_method",  22),
    ("9. แหล่งที่ได้มาซึ่งข้อมูล",                     "จากเจ้าของผู้ควบคุมโดยตรง",                            "data_source_direct",       22),
    (None,                                             "จากแหล่งอื่น",                                         "data_source_other",        22),
    ("10. ฐานในการประมวลผล",                           None,                                                   "legal_basis_thai",         28),
    ("11. ส่งหรือโอนข้อมูลไปต่างประเทศ",               "มีการส่งหรือโอนหรือไม่",                               "cross_border_transfer",    18),
    (None,                                             "บริษัทในเครือ",                                        "cross_border_affiliate",   22),
    (None,                                             "วิธีการโอนข้อมูล",                                     "cross_border_method",      22),
    (None,                                             "มาตรฐานการคุ้มครองข้อมูล",                             "cross_border_standard",    22),
    (None,                                             "ข้อยกเว้นตามมาตรา 28",                                 "cross_border_exception",   28),
    ("12. นโยบายการเก็บรักษาข้อมูล",                   "ประเภทของข้อมูลที่จัดเก็บ",                            "storage_type",             18),
    (None,                                             "วิธีการเก็บรักษา",                                     "storage_method",           22),
    (None,                                             "ระยะเวลาการเก็บรักษา",                                 "retention_period",         18),
    (None,                                             "สิทธิและวิธีการเข้าถึงข้อมูล",                         "access_rights",            28),
    (None,                                             "วิธีการลบหรือทำลายข้อมูล",                             "deletion_method",          22),
    ("13. มาตรการรักษาความมั่นคงปลอดภัย",              "มาตรการเชิงองค์กร",                                    "security_organizational",  22),
    (None,                                             "มาตรการเชิงเทคนิค",                                    "security_technical",       22),
    (None,                                             "มาตรการทางกายภาพ",                                     "security_physical",        22),
    (None,                                             "การควบคุมการเข้าถึงข้อมูล",                            "security_access_control",  22),
    (None,                                             "การกำหนดหน้าที่ความรับผิดชอบ",                         "security_responsibility",  22),
    (None,                                             "มาตรการตรวจสอบย้อนหลัง",                               "security_audit",           22),
]


def _get_value(record: RopaRecord, field: str) -> str | None:
    if field == "__seq__":
        return None  # filled separately
    if field == "data_subject_categories":
        return ", ".join(ds.name for ds in record.data_subjects) if record.data_subjects else None
    if field == "personal_data_types":
        return ", ".join(pdt.name for pdt in record.personal_data_types) if record.personal_data_types else None
    if field == "processor_name":
        return record.processor.name if record.processor else None
    if field == "source_controller":
        if record.processor and record.processor.source_controller:
            return record.processor.source_controller.name
        return None
    if field == "data_category":
        return record.processor.data_category if record.processor else None
    if field == "cross_border_transfer":
        val = getattr(record, field, None)
        if val is True:  return "มี"
        if val is False: return "ไม่มี"
        return None
    if field in ("retention_expiry_date", "next_review_date"):
        val = getattr(record, field, None)
        return val.isoformat() if val else None
    return getattr(record, field, None)


def _write_sheet(ws, records: list[RopaRecord], cols: list, sheet_title: str):
    """Write 2-row merged headers + data rows."""
    num_cols = len(cols)

    # ── Row 1: title banner ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    title_cell = ws.cell(row=1, column=1, value=sheet_title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    _style(title_cell, font=Font(bold=True, color="1F3864", name="TH SarabunPSK", size=14),
           fill=LIGHT_BLUE, alignment=CENTER)

    # ── Rows 2-3: group headers + sub-headers ────────────────────────────────
    ws.row_dimensions[2].height = 52
    ws.row_dimensions[3].height = 52

    group_start = None
    group_label = None

    for ci, (grp, sub, field, width) in enumerate(cols, start=1):
        _set_col_width(ws, ci, width)

        if grp is not None:
            # Close previous group merge if needed
            if group_start is not None and group_start < ci - 1:
                ws.merge_cells(start_row=2, start_column=group_start,
                               end_row=2, end_column=ci - 1)
            group_start = ci
            group_label = grp

        if sub is None:
            # Single-row header — merge rows 2-3
            cell2 = ws.cell(row=2, column=ci, value=group_label if grp is not None else "")
            ws.merge_cells(start_row=2, start_column=ci, end_row=3, end_column=ci)
            _style(cell2, font=WHITE_BOLD, fill=DARK_BLUE, alignment=CENTER, border=THIN_BORDER)
        else:
            # Group header in row 2, sub-header in row 3
            if grp is not None:
                cell2 = ws.cell(row=2, column=ci, value=group_label)
                _style(cell2, font=WHITE_BOLD, fill=DARK_BLUE, alignment=CENTER, border=THIN_BORDER)
            cell3 = ws.cell(row=3, column=ci, value=sub)
            _style(cell3, font=WHITE_BOLD, fill=MED_BLUE, alignment=CENTER, border=THIN_BORDER)

    # Close last group merge
    if group_start is not None and group_start < num_cols:
        # find last col with same group
        last = num_cols
        for ci in range(group_start + 1, num_cols + 1):
            if cols[ci - 1][0] is not None and ci > group_start:
                last = ci - 1
                break
        # merge row-2 cells for groups that have sub-headers
        # (already handled per-cell above; just ensure row-2 group cells are merged)

    # Re-pass to merge row-2 group cells across their sub-columns
    ci = 1
    while ci <= num_cols:
        grp, sub, _, _ = cols[ci - 1]
        if grp is not None and sub is not None:
            # find how many consecutive cols share this group (grp=None means continuation)
            end = ci
            while end + 1 <= num_cols and cols[end][0] is None:
                end += 1
            if end > ci:
                ws.merge_cells(start_row=2, start_column=ci, end_row=2, end_column=end)
                # re-style merged cell
                cell2 = ws.cell(row=2, column=ci)
                _style(cell2, font=WHITE_BOLD, fill=DARK_BLUE, alignment=CENTER, border=THIN_BORDER)
            ci = end + 1
        else:
            ci += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    for ri, record in enumerate(records, start=1):
        row = 3 + ri
        ws.row_dimensions[row].height = 18
        for ci, (_, _, field, _) in enumerate(cols, start=1):
            if field == "__seq__":
                val = ri
            else:
                val = _get_value(record, field)
            cell = ws.cell(row=row, column=ci, value=val)
            _style(cell, font=NORMAL_FONT, alignment=LEFT, border=THIN_BORDER)
            if ci == 1:  # seq column — center
                cell.alignment = CENTER

    # Freeze panes below headers
    ws.freeze_panes = ws.cell(row=4, column=2)


def export_excel(
    db: Session,
    user_id: int,
    search: str | None = None,
    department_id: int | None = None,
    role_type: str | None = None,
    risk_level: str | None = None,
    status: str | None = None,
) -> bytes:
    """Generate Excel with Controller and Processor sheets matching official ROPA format."""

    query = (
        db.query(RopaRecord)
        .filter(RopaRecord.is_deleted == False)
        .options(
            joinedload(RopaRecord.department),
            joinedload(RopaRecord.controller),
            joinedload(RopaRecord.processor).joinedload(Processor.source_controller),
            joinedload(RopaRecord.data_subjects),
            joinedload(RopaRecord.personal_data_types),
        )
    )

    # Status filter — default to approved only
    if status:
        query = query.filter(RopaRecord.status == status)
    else:
        query = query.filter(RopaRecord.status == "approved")

    if department_id:
        query = query.filter(RopaRecord.department_id == department_id)
    if role_type:
        query = query.filter(RopaRecord.role_type == role_type)
    if risk_level:
        query = query.filter(RopaRecord.risk_level == risk_level)
    if search:
        query = query.filter(RopaRecord.activity_name.ilike(f"%{search}%"))

    records = query.all()
    controller_records = [r for r in records if r.role_type == "Controller"]
    processor_records  = [r for r in records if r.role_type == "Processor"]

    wb = Workbook()

    ws_ctrl = wb.active
    ws_ctrl.title = "Controller"
    _write_sheet(ws_ctrl, controller_records, CONTROLLER_COLS, "บันทึกกิจกรรมการประมวลผลข้อมูลส่วนบุคคล — ผู้ควบคุมข้อมูล (Controller)")

    ws_proc = wb.create_sheet(title="Processor")
    _write_sheet(ws_proc, processor_records, PROCESSOR_COLS, "บันทึกกิจกรรมการประมวลผลข้อมูลส่วนบุคคล — ผู้ประมวลผลข้อมูล (Processor)")

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    file_bytes = output.read()

    log_action(
        db, user_id=user_id, action="export", table_name="ropa_records", record_id=0,
        new_value={
            "controller_count": len(controller_records),
            "processor_count": len(processor_records),
            "total": len(records),
            "filters": {"status": status or "approved", "department_id": department_id,
                        "role_type": role_type, "risk_level": risk_level, "search": search},
        },
        reason=f"Exported {len(records)} approved ROPA records to Excel",
    )

    return file_bytes
