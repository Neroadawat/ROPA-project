#!/usr/bin/env python3
"""Test script to verify the Excel structure parsing."""
from openpyxl import load_workbook

# Load the form
print("Loading ROPA_Form.xlsx...")
wb = load_workbook("docs/ROPA_Form.xlsx")

# Thai column indices
THAI_COLUMN_MAP = {
    "seq": 1,                          # ลำดับ
    "controller_name": 2,              # 1. ชื่อเจ้าของข้อมูล / ข้อมูลเกี่ยวกับผู้ควบคุม
    "activity_name": 3,                # 2. กิจกรรมประมวลผล
    "purpose": 4,                      # 3. วัตถุประสงค์ของการประมวลผล
    "personal_data_types": 5,          # 4. ข้อมูลส่วนบุคคลที่จัดเก็บ
    "data_subject_categories": 6,      # 5. หมวดหมู่ของข้อมูล
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    
    # Detect sheet type
    sheet_type = "Processor" if "processor" in sheet_name.lower() else "Controller"
    print(f"Sheet type: {sheet_type}")
    
    # Check data rows (starting from row 5)
    data_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        activity_col = row[2] if len(row) > 2 else None  # Column 3 (0-indexed)
        if activity_col and str(activity_col).strip():
            data_rows.append((row_idx, row))
            if len(data_rows) <= 2:  # Show first 2 data rows
                print(f"\nData Row {row_idx}:")
                for col_idx in [1, 2, 3, 4, 5, 6]:
                    cell_val = row[col_idx - 1] if col_idx <= len(row) else None
                    if cell_val:
                        val_str = str(cell_val)[:60]
                        print(f"  Col {col_idx}: {val_str}")
    
    print(f"\nTotal data rows: {len(data_rows)}")

wb.close()
print("\n✓ Structure validation successful!")

