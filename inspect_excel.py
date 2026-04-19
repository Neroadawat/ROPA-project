from openpyxl import load_workbook

wb = load_workbook(r'docs/ROPA_Form.xlsx')
ws = wb['Processor']

print('=== PROCESSOR SHEET - LOOKING FOR DATA ===\n')

# Show all rows with content
for row_idx in range(1, 30):
    row = ws[row_idx]
    has_content = False
    values = []
    for i, cell in enumerate(row[:15], 1):
        if cell.value:
            has_content = True
            val = str(cell.value)[:60] if cell.value else ''
            values.append(f"Col{i}: {val}")
    
    if has_content:
        print(f'\nRow {row_idx}:')
        for v in values:
            print(f"  {v}")

